from openpyxl import Workbook
from openpyxl.styles import Alignment

mywb = Workbook()
mysheet = mywb.active

mysheet['B2'] = 'cells merged together.'
mysheet.merge_cells('B2:D3')

mysheet.merge_cells('F6:F7')

cell = mysheet['F6']
cell.value = 'Two merged cells.'
cell.alignment = Alignment(horizontal='center',
                           vertical='center')

mywb.save('merging_cells.xlsx')
